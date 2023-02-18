#链家合肥二手房数据爬取
from openpyxl import load_workbook,Workbook
import requests
import pandas as pd
import openpyxl as op
from bs4 import BeautifulSoup
import numpy as np
hp=[]       #存放房屋价格
hb=[]       #存放房屋基本信息
hl=[]       #存放房屋地理位置

#爬取150页的数据
for i in range(1,151,1):
    #伪装请求头
    headers = {'User-Agent':'Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:95.0) Gecko/20100101 Firefox/95.0'}
    res=requests.get('https://hf.lianjia.com/ershoufang/pg'+str(i)+'/',headers=headers)
    house_info=BeautifulSoup(res.text,'html.parser')
    #爬取房价（how much/m**2)
    house_price=house_info.find_all('div',class_='unitPrice')
    # 爬取基本信息
    house_basic=house_info.find_all('div',class_="houseInfo")
    #爬取二手房地址
    house_loc=house_info.find_all('div',class_="positionInfo")
    for i in house_price:
        hp.append(i.text)
    for j in house_basic:
        hb.append(j.text)
    for k in house_loc:
        hl.append(k.text)


# 存储房屋相关信息
bg = op.load_workbook(r"data_get.xlsx")      	
sheet = bg.active    
for i in range(len(hb)):
    tmp=hb[i].split(sep=" | ")
    if len(tmp)==6:
        tmp.insert(5,"?")
    sheet.cell(row=i+1, column = 1, value =tmp[0])
    sheet.cell(row=i+1, column = 2, value =tmp[1])
    sheet.cell(row=i+1, column = 3, value =tmp[2])
    sheet.cell(row=i+1, column = 4, value =tmp[3])
    sheet.cell(row=i+1, column = 5, value =tmp[4])
    sheet.cell(row=i+1, column = 6, value =tmp[5])
    sheet.cell(row=i+1, column = 7, value =tmp[6])
    sheet.cell(row=i+1, column = 8, value =hl[i])
    sheet.cell(row=i+1, column = 9, value =hp[i])
bg.save("data_get.xlsx")



